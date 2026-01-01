"""Document parsing and chunking.

Supported file types:
- PDF: Azure Document Intelligence (native or converted via Gotenberg)
- Excel/CSV: openpyxl/csv parsers
- Markdown: Text file reader
"""

import csv
import json
import logging
import os
import time
import asyncio
import chardet
from pathlib import Path
from typing import Any, Dict, List, Tuple

import tiktoken
import openpyxl

from clients import get_azure_client, get_doc_intel_client
from core.config import (
    SMALL_MODEL_NAME,
    PARSE_MAX_TOKENS,
    PARSE_OVERLAP_TOKENS,
    PARSE_TOKENIZER_ENCODING,
    AI_TEMPERATURE,
    TABLE_MAX_TOKENS_PER_CHUNK,
    TABLE_EMPTY_ROW_THRESHOLD,
    TABLE_MAX_ROWS_TO_SCAN,
    DEBUG_SAVE_PROMPTS,
)
from ai import INTAKE_MINI_PROMPT
from core import (
    Unit,
    Location,
    File,
    Slice,
    Chunk,
    Cell,
    Dimensions,
    Sheet,
    Parse,
    Meta,
    BoundingBox,
    Line,
)

DEBUG_DIR = Path(__file__).parent / "testing"
if DEBUG_SAVE_PROMPTS:
    DEBUG_DIR.mkdir(exist_ok=True)


def polygon_to_bounds(
    polygon: List[float], page_width: float, page_height: float
) -> BoundingBox:
    """Convert Azure Document Intelligence polygon to normalized bounding box."""
    left = (polygon[0] / page_width) * 100
    top = (polygon[1] / page_height) * 100
    width = ((polygon[2] - polygon[0]) / page_width) * 100
    height = ((polygon[5] - polygon[1]) / page_height) * 100

    return BoundingBox(
        left=max(0, left),
        top=max(0, top),
        width=max(0, width),
        height=max(0, height),
    )


class Parser:
    """Document parser with chunking and metadata extraction."""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.tokenizer = tiktoken.get_encoding(PARSE_TOKENIZER_ENCODING)
        self.azure_client = get_azure_client()
        self.doc_intel_client = get_doc_intel_client()

    async def parse_document(self, file_path: str) -> list:
        """Parse document into structured data."""
        ext = Path(file_path).suffix.lower()

        if ext == ".pdf":
            return await self._parse_pdf(file_path)

        if ext in {".xlsx", ".xls"}:
            return await asyncio.to_thread(self._parse_excel, file_path)

        if ext == ".csv":
            return await asyncio.to_thread(self._parse_csv, file_path)

        if ext == ".md":
            return await asyncio.to_thread(self._parse_markdown, file_path)

        raise ValueError(f"Unsupported: {ext}")

    def get_intake_content(self, data: list, is_table: bool) -> str:
        """Extract text preview for metadata analysis."""
        if not data:
            raise ValueError("No data to extract content from")

        if is_table:
            return data[0]["text"]

        return "\n".join(line.text for line in data[0]["lines"])

    def get_full_text(self, data: list, is_table: bool) -> str:
        """Extract all text from parsed data."""
        if not data:
            raise ValueError("No data to extract text from")

        if is_table:
            return "\n\n".join(item["text"] for item in data)

        lines = []
        for page in data:
            for line in page["lines"]:
                lines.append(line.text)
        return "\n".join(lines)

    async def _parse_pdf(self, file_path: str) -> list:
        """Parse PDF using Azure Document Intelligence."""
        result = await self.doc_intel_client.analyze_document(file_path)

        page_data = []
        for page_idx, page in enumerate(result.pages, 1):
            if not page.lines:
                continue

            lines: List[Line] = []
            for line in page.lines:
                poly = line.polygon
                if poly and len(poly) >= 8:
                    bounds = polygon_to_bounds(poly, page.width, page.height)
                    lines.append(Line(text=line.content, bounds=bounds))
                else:
                    lines.append(Line(text=line.content))

            page_data.append({"page": page_idx, "lines": lines})

        return page_data

    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse Excel file into sheet data."""
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_data = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if not sheet.max_row:
                continue

            rows = []
            for row in sheet.iter_rows(max_row=min(sheet.max_row, TABLE_MAX_ROWS_TO_SCAN)):
                rows.append([cell.value for cell in row])

            max_row, max_col = self._get_table_bounds(rows)
            text, cells = self._build_table_text(rows, max_row, max_col)

            sheet_data.append({
                "metadata": {"sheet_name": sheet_name, "sheet_index": len(sheet_data) + 1},
                "text": text,
                "cells": cells,
                "dimensions": {"max_row": max_row, "max_col": max_col},
            })

        workbook.close()
        return sheet_data

    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse CSV file into sheet data."""
        with open(file_path, "rb") as f:
            sample = f.read(65536)

        encoding = chardet.detect(sample)["encoding"]
        delimiter = csv.Sniffer().sniff(sample.decode(encoding)[:4096]).delimiter

        rows = []
        with open(file_path, "r", encoding=encoding) as f:
            for row in csv.reader(f, delimiter=delimiter):
                rows.append(row)
                if len(rows) >= TABLE_MAX_ROWS_TO_SCAN:
                    break

        max_row, max_col = self._get_table_bounds(rows)
        text, cells = self._build_table_text(rows, max_row, max_col)

        return [{
            "metadata": {"sheet_name": "Data", "sheet_index": 1},
            "text": text,
            "cells": cells,
            "dimensions": {"max_row": max_row, "max_col": max_col},
        }]

    def _get_table_bounds(self, rows: List[List]) -> tuple[int, int]:
        """Find content boundaries."""
        max_row = 0
        max_col = 0
        empty_count = 0

        for row_idx, row in enumerate(rows, 1):
            has_content = False
            for col_idx, value in enumerate(row, 1):
                if value is not None and str(value).strip():
                    has_content = True
                    max_col = max(max_col, col_idx)

            if has_content:
                max_row = row_idx
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= TABLE_EMPTY_ROW_THRESHOLD:
                    break

        return max_row, max_col

    def _build_table_text(
        self, rows: List[List], max_row: int, max_col: int
    ) -> Tuple[str, Dict[str, Cell]]:
        """Build text representation and cells dict."""
        lines = []
        cells = {}

        for row_idx, row in enumerate(rows[:max_row], 1):
            row_values = []
            for col_idx in range(max_col):
                value = row[col_idx] if col_idx < len(row) else None
                col_letter = self._col_letter(col_idx + 1)
                coord = f"{col_letter}{row_idx}"

                if value is not None:
                    clean = str(value).replace("\n", " ").replace("\r", " ").strip()
                    row_values.append(clean)
                    if clean:
                        cells[coord] = Cell(value=clean, row=row_idx, col=col_letter)
                else:
                    row_values.append("")

            lines.append(" | ".join(row_values))

        return "\n".join(lines), cells

    def _col_letter(self, n: int) -> str:
        """Convert column number to letter (1=A, 27=AA)."""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + n % 26) + result
            n //= 26
        return result

    def _parse_markdown(self, file_path: str) -> list:
        """Parse markdown file into page_data."""
        with open(file_path, "r", encoding="utf-8") as f:
            text = f.read()

        lines = [Line(text=line) for line in text.split("\n") if line.strip()]
        return [{"page": 1, "lines": lines}]

    def build_chunks(self, data: list, file: File) -> Parse:
        """Build chunks from parsed data."""
        is_table = file.name.lower().endswith((".xlsx", ".xls", ".csv"))

        if is_table:
            result = self._build_table_chunks(data, file)
        else:
            result = self._build_text_chunks(data, file)

        if DEBUG_SAVE_PROMPTS:
            self._dump_debug(file.name, result)

        return result

    def _build_text_chunks(self, data: list, file: File) -> Parse:
        """Build chunks for text documents (PDF, Markdown)."""
        content: Dict[str, Unit] = {}
        all_units: List[Tuple[Unit, int]] = []

        unit_num = 1
        for page in data:
            page_num = page["page"]
            for line in page["lines"]:
                unit_id = str(unit_num)
                unit = Unit(
                    id=unit_id,
                    type="text",
                    text=line.text,
                    location=Location(page=page_num, bounds=line.bounds),
                )
                tokens = len(self.tokenizer.encode(line.text))
                all_units.append((unit, tokens))
                content[unit_id] = unit
                unit_num += 1

        if not all_units:
            return Parse(chunks=[], content={}, sheets={})

        # Build chunks with overlap
        chunks: List[Chunk] = []
        idx = 0
        while idx < len(all_units):
            chunk_units = []
            chunk_tokens = 0
            start_idx = idx

            while idx < len(all_units):
                unit, tokens = all_units[idx]
                if chunk_tokens + tokens > PARSE_MAX_TOKENS and chunk_units:
                    break
                chunk_units.append(unit)
                chunk_tokens += tokens
                idx += 1

            if chunk_units:
                chunks.append(Chunk(file=file, units=chunk_units, tokens=chunk_tokens))

            # Overlap
            if idx < len(all_units):
                overlap = 0
                back = idx
                while back > start_idx + 1:
                    back -= 1
                    _, t = all_units[back]
                    overlap += t
                    if overlap >= PARSE_OVERLAP_TOKENS:
                        break
                idx = back

        return Parse(chunks=chunks, content=content, sheets={})

    def _build_table_chunks(self, data: list, file: File) -> Parse:
        """Build chunks for table documents (Excel, CSV)."""
        content: Dict[str, Unit] = {}
        chunks: List[Chunk] = []
        sheets: Dict[str, Sheet] = {}

        for item in data:
            sheet_name = item["metadata"]["sheet_name"]
            sheet_text = item["text"]
            cells: Dict[str, Cell] = item["cells"]
            dims = item["dimensions"]

            units = []
            for coord, cell in cells.items():
                unit = Unit(
                    id=coord,
                    type="table",
                    text=cell.value,
                    location=Location(sheet=sheet_name, row=cell.row, col=cell.col),
                )
                units.append(unit)
                content[coord] = unit

            units.sort(key=lambda u: (u.location.row, u.location.col))
            sheet_tokens = len(self.tokenizer.encode(sheet_text))

            sheets[sheet_name] = Sheet(
                cells=cells,
                dimensions=Dimensions(max_row=dims["max_row"], max_col=dims["max_col"]),
                tokens=sheet_tokens,
            )

            if sheet_tokens > TABLE_MAX_TOKENS_PER_CHUNK:
                chunk_units, tokens = self._truncate_units(units, TABLE_MAX_TOKENS_PER_CHUNK)
                truncated = True
            else:
                chunk_units = units
                tokens = sheet_tokens
                truncated = False

            chunks.append(Chunk(
                file=file,
                units=chunk_units,
                tokens=tokens,
                slice=Slice(sheet=sheet_name, truncated=truncated),
            ))

        return Parse(chunks=chunks, content=content, sheets=sheets)

    def _truncate_units(self, units: List[Unit], max_tokens: int) -> Tuple[List[Unit], int]:
        """Truncate units to fit token budget."""
        result = []
        total = 0
        for unit in units:
            t = len(self.tokenizer.encode(unit.text))
            if total + t > max_tokens and result:
                break
            result.append(unit)
            total += t
        return result, total

    async def analyze_document_metadata(self, content: str, file_name: str) -> Meta:
        """Extract document metadata using AI."""
        try:
            preview = content[:2000] + "..." if len(content) > 2000 else content
            prompt = INTAKE_MINI_PROMPT.format(document_text=preview)

            response = await self.azure_client.chat.completions.create(
                model=SMALL_MODEL_NAME,
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": "Analyze this document."},
                ],
                temperature=AI_TEMPERATURE,
                response_format={"type": "json_object"},
            )

            data = json.loads(response.choices[0].message.content)
            return Meta(
                company=data.get("company"),
                ticker=data.get("ticker"),
                doc_type=data.get("doc_type"),
                period_label=data.get("period_label"),
                blurb=data.get("blurb"),
            )
        except Exception as e:
            self.logger.error(f"Metadata extraction failed: {e}")
            return Meta(doc_type="other", blurb=f"Document: {file_name}")

    def _dump_debug(self, file_name: str, result: Parse) -> None:
        """Dump result to JSON for debugging."""
        name = os.path.basename(file_name).replace(" ", "_")
        dump_file = DEBUG_DIR / f"{name}_{int(time.time())}.json"
        with open(dump_file, "w") as f:
            json.dump({"file": name, "result": result.model_dump()}, f, indent=2)
        self.logger.info(f"[DEBUG] Dumped to {dump_file}")
